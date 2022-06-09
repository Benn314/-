import openpyxl
from wordcloud import WordCloud

# 与即时显示图片相关的模块
'''
import matplotlib.pyplot as plt   # 绘制图像的模块
import numpy as np
from PIL import Image
'''

# 以省份的确诊病例总数代表其出现的频率
frequency_in = {}  # frequency_in表示国内的情况，frequency_out表示国外的情况
# 读取文件
wb = openpyxl.load_workbook('data.xlsx')
# 获取工作表
ws = wb['国内疫情']
# 对工作表进行遍历
for row in ws.values:
    # print(row)  # 以元组的形式反馈，只要通过序列就可以获取当中的值，返回结果第一列不需要，我们采用判断语句给它去掉
    if row[0] == '省份':
        pass
    else:
        frequency_in[row[0]] = float(row[1])  # 采用键值对来对应关联,跟map容器一样

# print(frequency)    # 我们用词云来绘制，所以这段输出语句注释掉
wordcloud = WordCloud(font_path="C:/Windows/Fonts/SIMLI.TTF",  # 这里可以指定字体样式,字体代码名称可以在相应的属性中查看
                      background_color="white",  # 这里可以指定背景颜色
                      width=1920, height=1080)  # 还可以指定输出图片的宽度和高度

# # 根据确诊病例的数目生成词云
# wordcloud.generate_from_frequencies(frequency_in)  # 这里报错TypeError: unsupported operand type(s) for /: 'str' and 'float'
# # 我们需要将传入的frequency改成字符串或浮点型
#
# # 保存词云
# wordcloud.to_file('wordcloud.png')

# 我们知道，不同国家是以洲为单位来划分的
frequency_out = {}
# print(wb.sheetnames)
sheet_name = wb.sheetnames
for each in sheet_name:
    if "洲" in each:
        ws = wb[each]
        for row in ws.values:
            # print(row)  #用键值对把国家和累计确诊的人数一一对应起来
            if row[0] == '国家':
                pass
            else:
                frequency_out[row[0]] = float(row[1])
    else:
        pass


# print(frequency_out)
# # 根据确诊病例的数目生成词云
# wordcloud.generate_from_frequencies(frequency_out)  # 这里报错TypeError: unsupported operand type(s) for /: 'str' and 'float'
# # 我们需要将传入的frequency改成字符串或浮点型
# #键值越大，键在词云图中生成就越大   字体越大，累计确诊人数越大，代表该地区疫情越严重
# # 保存词云
# wordcloud.to_file('wordcloud1.png')

# 由于绘制国内疫情词云图和世界疫情词云图有很多公共的部分，为了提高代码的复用性，我们封装一个方法来调用生成想要词云国
def generate_pic(frequency, name):
    # 这里可以事先准备一张图片，可以用作背景
    # background_image = np.array(Image.open('pic.jpg'))
    wordcloud = WordCloud(font_path="C:/Windows/Fonts/SIMLI.TTF",  # 这里可以指定字体样式,字体代码名称可以在相应的属性中查看
                          background_color="white",  # 这里可以指定背景颜色
                          width=1920, height=1080)  # 还可以指定输出图片的宽度和高度
    # 根据确诊病例的数目生成词云
    wordcloud.generate_from_frequencies(frequency)
    # 保存词云
    wordcloud.to_file('%s.png' % (name))


# 调用函数测试,在主函数调用就好
# generate_pic(frequency_in, '国内疫情词云图')
# generate_pic(frequency_out, '国外疫情词云图')

# 词云绘制完成，下一步我们采集各个省份以及地区的数据信息来绘制我们的地图，先新建 data_get.py 来采集数据
